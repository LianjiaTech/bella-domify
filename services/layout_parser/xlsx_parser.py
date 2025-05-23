# -*- coding: utf-8 -*-
# ===============================================================
#
#    Copyright (C) 2024 Beike, Inc. All Rights Reserved.
#
#    @Create Author : luxu(luxu002@ke.com)
#    @Create Time   : 2024/11/7
#    @Description   :
#
# ===============================================================
import os
from io import BytesIO

import openpyxl

from services.constants import TEXT
from services.layout_parse_utils import get_s3_links_for_simple_block_batch
from services.simple_block import SimpleBlock


def layout_parse(byte_data):
    # 将字节数据转换为BytesIO对象
    byte_stream = BytesIO(byte_data)

    # 打开XLSX文件
    workbook = openpyxl.load_workbook(byte_stream)

    # 用于存储所有内容的字符串
    all_content = ""

    # 遍历每个工作表
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        all_content += f"Sheet: {sheet}\n"

        # 遍历每一行
        for row in worksheet.iter_rows():
            # 遍历每一列
            for cell in row:
                # 将单元格的值拼接到字符串中
                all_content += str(cell.value) + "\t"
            all_content += "\n"
        all_content += "\n"

    result_text = all_content
    result_json = get_s3_links_for_simple_block_batch([SimpleBlock(type=TEXT, text=result_text)])
    return result_json, result_text


def markdown_parse(byte_data):
    # 将字节数据转换为BytesIO对象
    byte_stream = BytesIO(byte_data)

    # 打开XLSX文件
    workbook = openpyxl.load_workbook(byte_stream)

    # 用于存储所有内容的字符串
    all_content = ""

    # 遍历每个工作表
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        all_content += f"## Sheet: {sheet}\n\n"  # 添加工作表标题

        # 用于存储当前工作表的Markdown表格内容
        markdown_table = []

        # 遍历每一行
        for row in worksheet.iter_rows(values_only=True):
            # 将每一行的内容转换为Markdown表格行
            markdown_row = "| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |"
            markdown_table.append(markdown_row)

            # 如果是第一行，添加表头分隔线
            if len(markdown_table) == 1:
                markdown_table.append("| " + " | ".join(["---"] * len(row)) + " |")

        # 将当前工作表的Markdown表格内容添加到总内容中
        all_content += "\n".join(markdown_table) + "\n\n"

    return all_content


if __name__ == "__main__":

    file_path = os.path.abspath(__file__).split("document_parse")[0] + "document_parse/test/samples/file_type_demo/"

    file_name = 'demo.xlsx'

    # 读取本地文件
    try:
        with open(file_path + file_name, 'rb') as file:
            buf_data = file.read()
    except Exception as e:
        print(f"读取文件失败: {e}")

    print(layout_parse(buf_data))
